# -*- coding: utf-8 -*-
"""
数据导出模块 — CSV / XLSX / ZIP 批量导出
V3.7.2 — 日期筛选 / 字段选择 / Excel 格式 / 多表打包
"""
from typing import Optional, List
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import inspect
from backend.database import get_db
from pydantic import BaseModel
import io, csv, zipfile
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _safe_filename(name: str) -> str:
    """确保文件名只有 ASCII 字符"""
    safe = name.encode('ascii', 'ignore').decode('ascii').strip()
    return safe or 'export'

router = APIRouter(prefix="/api/export", tags=["数据导出"])

# ═══════════════════════════════════════════
# 配置
# ═══════════════════════════════════════════

EXPORT_TABLES = {
    "members": "会员信息", "staff": "员工信息", "courses": "课程信息",
    "sales": "售课记录", "class_records": "上课记录", "checkins": "进场记录",
    "wristbands": "手环管理", "body_measurements": "体测记录", "recharges": "充值记录",
    "membership_cards": "会籍卡管理", "products": "商品管理", "product_sales": "商品零售",
    "finance_income": "收入记录", "finance_expense": "支出记录", "operation_logs": "操作日志",
}

# 每个表的主日期列（用于日期范围筛选）
DATE_COLUMNS = {
    "members": ["signup_date", "end_date", "birth_date"],
    "staff": ["hire_date", "birth_date"],
    "courses": [],
    "sales": ["sale_date"],
    "class_records": ["class_date"],
    "checkins": ["checkin_date"],
    "wristbands": [],
    "body_measurements": ["measure_date"],
    "recharges": ["recharge_date"],
    "membership_cards": ["start_date", "end_date"],
    "products": [],
    "product_sales": ["sale_date"],
    "finance_income": ["income_date"],
    "finance_expense": ["expense_date"],
    "operation_logs": ["created_at"],
}

CN_NAMES = {
    "member_id": "会员编号", "member_name": "会员姓名",
    "staff_id": "员工编号", "course_id": "课程编号", "sale_id": "售课编号",
    "record_id": "记录编号", "card_id": "卡号", "band_id": "手环编号",
    "product_id": "商品编号", "sale_id": "零售编号",
    "name": "姓名", "gender": "性别", "phone": "手机号",
    "member_type": "会员类型", "source": "来源", "level": "等级",
    "status": "状态", "balance": "余额",
    "position": "岗位", "base_salary": "底薪", "sale_commission_rate": "售课提成",
    "course_name": "课程名", "course_type": "课程类型", "sport_type": "运动类型",
    "standard_hours": "标准课时", "standard_price": "标准价", "discount_price": "优惠价",
    "sale_date": "售课日期", "class_date": "上课日期", "checkin_date": "进场日期",
    "unit_price": "单价", "total_price": "总价", "actual_amount": "实收金额",
    "income_date": "收入日期", "expense_date": "支出日期",
    "amount": "金额", "category": "类别", "payment_method": "支付方式", "payee": "收款方",
    "height": "身高", "weight": "体重", "body_fat": "体脂率",
    "muscle_mass": "肌肉量", "bmi": "BMI", "basal_metabolism": "基础代谢",
    "recharge_amount": "充值金额", "bonus_amount": "赠送金额",
    "card_type": "卡类型", "duration_days": "有效期", "price": "售价",
    "start_date": "开始日期", "end_date": "截止日期",
    "cost_price": "进价", "selling_price": "售价",
    "stock": "库存", "unit": "单位", "supplier": "供应商", "min_stock": "安全库存",
    "product_name": "商品名", "quantity": "数量",
    "username": "操作人", "action": "操作", "resource": "资源",
    "resource_id": "资源ID", "detail": "详情", "ip_address": "IP地址",
    "created_at": "创建时间",
    "hire_date": "入职日期", "birth_date": "出生日期",
    "signup_date": "注册日期", "measure_date": "测量日期",
    "recharge_date": "充值日期", "remind_date": "提醒日期",
    "valid_days": "有效期", "max_bookings": "最大预约数",
    "coach": "教练", "location": "地点",
    "id_card": "身份证", "bank_card": "银行卡",
    "remark": "备注", "operator": "操作人",
    "pass_id": "月卡编号", "pass_name": "月卡名称", "pass_type": "月卡类型",
    "purchase_date": "购买日期", "valid_from": "有效期起", "valid_until": "有效期止",
    "included_courses": "包含课程", "course_names": "课程名称",
    "package_id": "课程包编号", "package_name": "课程包名称", "package_type": "课程包类型",
    "total_count": "总次数", "used_hours": "已用课时", "total_hours": "总课时",
    "remaining_hours": "剩余课时",
    "inbound_id": "入库编号", "inbound_date": "入库日期",
    "unit_cost": "进货单价", "total_cost": "总成本",
    "store_id": "门店编号",
}


# ═══════════════════════════════════════════
# 模型映射
# ═══════════════════════════════════════════

def _get_model_map():
    from backend.models.models import (
        Member, Staff, Course, Sale, ClassRecord, Checkin, Wristband,
        BodyMeasurement, Recharge, MembershipCard, Product, ProductSale,
        FinanceIncome, FinanceExpense,
    )
    from backend.routers.operation_log import OperationLog
    return {
        "members": Member, "staff": Staff, "courses": Course,
        "sales": Sale, "class_records": ClassRecord, "checkins": Checkin,
        "wristbands": Wristband, "body_measurements": BodyMeasurement,
        "recharges": Recharge, "membership_cards": MembershipCard,
        "products": Product, "product_sales": ProductSale,
        "finance_income": FinanceIncome, "finance_expense": FinanceExpense,
        "operation_logs": OperationLog,
    }


def _get_model(table_name: str):
    model_map = _get_model_map()
    model = model_map.get(table_name)
    if not model:
        raise HTTPException(404, f"表「{table_name}」不存在")
    return model


# ═══════════════════════════════════════════
# 辅助：列信息
# ═══════════════════════════════════════════

def _get_column_type(model, col_key: str) -> str:
    """推断列的类型：string / number / date"""
    try:
        mapper = inspect(model)
        attr = getattr(mapper.attrs, col_key, None)
        if attr is None:
            return "string"
        col_type = str(attr.expression.type).lower()
        if "int" in col_type or "float" in col_type or "decimal" in col_type or "numeric" in col_type:
            return "number"
        if "date" in col_type or "time" in col_type or "datetime" in col_type:
            return "date"
    except Exception:
        pass
    return "string"


def _get_columns(model, skip_id=True):
    """获取模型的所有列名"""
    mapper = inspect(model)
    return [c.key for c in mapper.attrs if not (skip_id and c.key == 'id')]


# ═══════════════════════════════════════════
# 导出格式生成器
# ═══════════════════════════════════════════

def _make_cn_header(cols: list) -> list:
    return [CN_NAMES.get(c, c) for c in cols]


def _row_values(row, cols: list) -> list:
    vals = []
    for c in cols:
        val = getattr(row, c, "")
        if val is None:
            val = ""
        if hasattr(val, 'isoformat'):
            val = val.isoformat()
        vals.append(str(val))
    return vals


def _generate_csv(rows, cols: list) -> io.StringIO:
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(_make_cn_header(cols))
    for row in rows:
        w.writerow(_row_values(row, cols))
    output.seek(0)
    return output


def _generate_xlsx(rows, cols: list, sheet_name: str = "Sheet1") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = _make_cn_header(cols)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 数据行
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(cols, 1):
            val = getattr(row, col, "")
            if val is None:
                val = ""
            if hasattr(val, 'isoformat'):
                val = val.isoformat()
            ws.cell(row=ri, column=ci, value=str(val))

    # 自适应列宽
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + ci) if ci <= 26 else 'A'].width = max(12, min(len(h) * 2 + 4, 40))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ═══════════════════════════════════════════
# 查询构建
# ═══════════════════════════════════════════

def _build_query(model, db: Session, date_col: str = "", date_from: str = "", date_to: str = ""):
    q = db.query(model)
    if date_col and hasattr(model, date_col):
        col = getattr(model, date_col)
        if date_from:
            try:
                q = q.filter(col >= datetime.strptime(date_from, "%Y-%m-%d").date())
            except ValueError:
                pass
        if date_to:
            try:
                q = q.filter(col <= datetime.strptime(date_to, "%Y-%m-%d").date())
            except ValueError:
                pass
    return q.order_by(model.id)


# ═══════════════════════════════════════════
# 端点
# ═══════════════════════════════════════════

@router.get("/tables")
def list_tables():
    """列出所有可导出表"""
    return [{"key": k, "name": v} for k, v in EXPORT_TABLES.items()]


@router.get("/{table_name}/fields")
def list_fields(table_name: str):
    """列出指定表的可用字段（含中文名和类型）"""
    model = _get_model(table_name)
    cols = _get_columns(model)
    return [{
        "key": c,
        "name": CN_NAMES.get(c, c),
        "type": _get_column_type(model, c),
    } for c in cols]


@router.get("/{table_name}")
def export_table(
    table_name: str,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    fields: str = Query("", description="逗号分隔的字段名"),
    date_col: str = Query("", description="日期筛选列"),
    date_from: str = Query("", description="开始日期 YYYY-MM-DD"),
    date_to: str = Query("", description="截止日期 YYYY-MM-DD"),
    db: Session = Depends(get_db),
):
    """导出指定表（支持字段选择、日期范围、CSV/XLSX）"""
    model = _get_model(table_name)
    all_cols = _get_columns(model)

    # 字段选择
    if fields:
        selected = [f.strip() for f in fields.split(",") if f.strip()]
        cols = [c for c in all_cols if c in selected] or all_cols
    else:
        cols = all_cols

    # 查询
    rows = _build_query(model, db, date_col, date_from, date_to).all()

    safe_name = table_name
    date_str = datetime.now().strftime("%Y%m%d")

    if format == "xlsx":
        buf = _generate_xlsx(rows, cols, sheet_name=safe_name)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_name}_{date_str}.xlsx"',
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
        )
    else:
        output = _generate_csv(rows, cols)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv; charset=utf-8-sig",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_name}_{date_str}.csv"',
                "Content-Type": "text/csv; charset=utf-8-sig",
            },
        )


class BatchExportRequest(BaseModel):
    tables: List[str]
    format: str = "csv"
    date_from: str = ""
    date_to: str = ""


@router.post("/batch")
def batch_export(data: BatchExportRequest, db: Session = Depends(get_db)):
    """批量导出多表为 ZIP 压缩包"""
    if not data.tables:
        raise HTTPException(400, "请选择至少一个表")

    zip_buf = io.BytesIO()
    date_str = datetime.now().strftime("%Y%m%d")

    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for tname in data.tables:
            if tname not in EXPORT_TABLES:
                continue
            safe_name = tname
            try:
                model = _get_model(tname)
                cols = _get_columns(model)
                date_cols = DATE_COLUMNS.get(tname, [])
                dc = date_cols[0] if date_cols else ""
                rows = _build_query(model, db, dc, data.date_from, data.date_to).all()
                if data.format == "xlsx":
                    buf = _generate_xlsx(rows, cols, sheet_name=safe_name)
                    ext = "xlsx"
                else:
                    output = _generate_csv(rows, cols)
                    buf = io.BytesIO(output.getvalue().encode("utf-8-sig"))
                    ext = "csv"

                zf.writestr(f"{safe_name}_{date_str}.{ext}", buf.getvalue())
            except Exception as e:
                zf.writestr(f"{safe_name}_ERROR.txt", str(e))

    zip_buf.seek(0)
    return StreamingResponse(
        iter([zip_buf.getvalue()]),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="export_batch_{date_str}.zip"',
            "Content-Type": "application/zip",
        },
    )
