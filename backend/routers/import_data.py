# -*- coding: utf-8 -*-
"""
数据导入模块 — 标准模板导入 + 进度跟踪
V3.8.2
"""
from typing import Optional, List, Tuple
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import HTMLResponse, Response, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from backend.database import get_db, SessionLocal
from backend.models.models import ImportTask, Member, Staff, Course, MembershipCard, Sale, Checkin, BodyMeasurement
from backend.services.id_gen import generate_id
from backend.routers.export_data import CN_NAMES
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io, json, os, threading, asyncio
from datetime import date, datetime
from decimal import Decimal

router = APIRouter(prefix="/api/import", tags=["数据导入"])

# ═══════════════════════════════════════════
# 导入类型配置
# ═══════════════════════════════════════════

# (field, chinese_header, required)
IMPORT_FIELDS = {
    "member": [
        ("name", "姓名*", True),
        ("phone", "手机号*", True),
        ("gender", "性别", False),
        ("birth_date", "出生日期", False),
        ("height", "身高(cm)", False),
        ("weight", "体重(kg)", False),
        ("body_fat", "体脂率(%)", False),
        ("level", "会员等级", False),
        ("status", "状态", False),
        ("source", "客户来源", False),
        ("staff_id", "跟进员工编号", False),
        ("remark", "备注", False),
        ("id_card", "身份证号", False),
    ],
    "staff": [
        ("name", "姓名*", True),
        ("phone", "手机号*", True),
        ("gender", "性别", False),
        ("birth_date", "出生日期", False),
        ("position", "岗位", False),
        ("hire_date", "入职日期", False),
        ("base_salary", "底薪", False),
        ("status", "状态", False),
        ("id_card", "身份证号", False),
        ("bank_card", "银行卡号", False),
        ("sale_commission_rate", "售课提成比例(%)", False),
        ("class_commission_rate", "上课提成比例(%)", False),
    ],
    "membership_card": [
        ("member_phone", "会员手机号*", True),
        ("member_name", "会员姓名*", True),
        ("card_type", "会籍类型", False),
        ("card_name", "卡名称", False),
        ("duration_days", "有效期(天)", False),
        ("price", "售价", False),
        ("actual_amount", "实收金额*", True),
        ("start_date", "有效期起", False),
        ("end_date", "有效期止", False),
        ("total_classes", "购买次数", False),
        ("bonus_classes", "赠送次数", False),
        ("face_value", "面值/余额总额", False),
        ("consumed_amount", "已消费金额", False),
        ("status", "状态", False),
        ("remark", "备注", False),
        ("staff_phone", "销售员手机号", False),
    ],
    "sold_lesson": [
        ("member_phone", "会员手机号*", True),
        ("member_name", "会员姓名*", True),
        ("course_name", "课程名称", False),
        ("course_type", "课程种类", False),
        ("sale_date", "售课日期", False),
        ("bought_hours", "购买课时", False),
        ("bonus_hours", "赠送课时", False),
        ("total_hours", "总课时", False),
        ("unit_price", "单价", False),
        ("total_price", "总价", False),
        ("actual_amount", "实收金额", False),
        ("payment_method", "付款方式", False),
        ("staff_phone", "销售员手机号", False),
        ("remark", "备注", False),
    ],
    "course": [
        ("name", "课程名称*", True),
        ("sport_type", "运动项目", False),
        ("course_type", "课程类型", False),
        ("standard_hours", "标准课时", False),
        ("standard_price", "标准售价", False),
        ("discount_price", "优惠售价", False),
        ("valid_days", "有效期(天)", False),
        ("status", "状态", False),
        ("max_bookings", "最大预约人数", False),
        ("coach", "教练", False),
    ],
    "checkin": [
        ("member_phone", "会员手机号*", True),
        ("member_name", "会员姓名*", True),
        ("checkin_date", "进场日期*", True),
        ("checkin_time", "进场时间", False),
        ("checkin_type", "进场类型", False),
        ("card_type", "卡片类型", False),
        ("card_id", "核销会籍卡编号", False),
        ("consume_type", "核销方式", False),
        ("consume_detail", "核销明细", False),
        ("operator", "操作员", False),
    ],
    "body_measurement": [
        ("member_phone", "会员手机号*", True),
        ("member_name", "会员姓名*", True),
        ("measure_date", "体测日期*", True),
        ("height", "身高(cm)", False),
        ("weight", "体重(kg)", False),
        ("body_fat", "体脂率(%)", False),
        ("bmi", "BMI", False),
        ("muscle_mass", "肌肉量(kg)", False),
        ("basal_metabolism", "基础代谢(kcal)", False),
        ("body_age", "体年龄", False),
    ],
}

# Chinese header → field mapping per type
def _build_header_map(import_type: str) -> dict:
    """构建中文表头 → 字段名映射"""
    return {hdr: fld for fld, hdr, _ in IMPORT_FIELDS[import_type]}

# Type config
TYPE_CONFIG = {
    "member": {
        "model": Member,
        "id_prefix": "M",
        "key_field": None,  # None = 严格校验后新建（不允许重复）
        "name": "会员",
    },
    "staff": {
        "model": Staff,
        "id_prefix": "S",
        "key_field": "phone",
        "name": "员工",
    },
    "membership_card": {
        "model": MembershipCard,
        "id_prefix": "MC",
        "key_field": None,  # 始终 INSERT
        "name": "已售会籍卡",
    },
    "sold_lesson": {
        "model": Sale,
        "id_prefix": "SL",
        "key_field": None,  # 始终 INSERT
        "name": "已售私教课",
    },
    "course": {
        "model": Course,
        "id_prefix": "CO",
        "key_field": "name",
        "name": "课程",
    },
    "checkin": {
        "model": Checkin,
        "id_prefix": "CI",
        "key_field": None,
        "name": "进场记录",
    },
    "body_measurement": {
        "model": BodyMeasurement,
        "id_prefix": "BM",
        "key_field": None,
        "name": "体测记录",
    },
}

# Date fields per type (these need date parsing)
DATE_FIELDS = {
    "member": {"birth_date"},
    "staff": {"birth_date", "hire_date"},
    "membership_card": {"start_date", "end_date"},
    "sold_lesson": {"sale_date"},
    "course": set(),
    "checkin": {"checkin_date"},
    "body_measurement": {"measure_date"},
}

# Decimal fields per type
DECIMAL_FIELDS = {
    "member": {"height", "weight", "body_fat"},
    "staff": {"base_salary", "sale_commission_rate", "class_commission_rate"},
    "membership_card": {"price", "face_value", "consumed_amount", "actual_amount"},
    "sold_lesson": {"unit_price", "total_price", "actual_amount"},
    "course": {"standard_price", "discount_price"},
    "checkin": set(),
    "body_measurement": {"height", "weight", "body_fat", "bmi", "muscle_mass", "basal_metabolism", "body_age"},
}

# Protected fields per type (never overwritten during upsert)
PROTECTED_FIELDS = {
    "member": {"member_id", "balance", "recharge_total", "consumed_amount",
               "total_lessons", "used_lessons", "remaining_lessons",
               "total_checkin_days", "last_checkin_date", "photo_path",
               "created_at", "updated_at"},
    "staff": {"staff_id", "sale_commission_amount", "class_commission_amount",
              "total_commission", "month_sale_amount", "month_class_count",
              "month_sale_commission", "month_class_commission", "month_total_commission",
              "today_class_count", "created_at", "updated_at"},
    "membership_card": {"id", "card_id", "member_id", "is_product",
                        "voided", "void_reason", "void_time", "void_operator", "created_at"},
    "sold_lesson": {"id", "sale_id", "member_id", "member_name", "member_phone",
                    "commission_rate", "commission_amount", "operator",
                    "deposit", "discount", "payment_status", "balance_due",
                    "voided", "void_reason", "void_time", "void_operator", "created_at"},
    "course": {"id", "course_id", "created_at", "updated_at"},
    "checkin": {"id", "checkin_id", "member_id", "created_at"},
    "body_measurement": {"id", "measure_id", "member_id", "created_at"},
}


# ═══════════════════════════════════════════
# 辅助函数
# ═══════════════════════════════════════════

def _parse_cell_value(val, field: str, import_type: str):
    """将 openpyxl 单元格值转换为合适的 Python 类型"""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    if field in DATE_FIELDS.get(import_type, set()):
        # Handle both datetime objects and strings
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        if isinstance(val, str):
            try:
                return date.fromisoformat(val.strip())
            except ValueError:
                return None
        return val
    if field in DECIMAL_FIELDS.get(import_type, set()):
        try:
            return Decimal(str(val))
        except (ValueError, TypeError):
            return Decimal(0)
    # String fields
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, (int, float)):
        return val
    return str(val)


def _type_name_cn(import_type: str) -> str:
    names = {"member": "会员", "staff": "员工", "membership_card": "会籍卡", "sold_lesson": "私教课",
             "course": "课程", "checkin": "进场记录", "body_measurement": "体测记录"}
    return names.get(import_type, import_type)


# ═══════════════════════════════════════════
# 模板下载 — GET /api/import/template/{type}
# ═══════════════════════════════════════════

@router.get("/template/{import_type}")
def download_template(import_type: str):
    """下载导入模板 XLSX"""
    if import_type not in IMPORT_FIELDS:
        raise HTTPException(404, f"不支持的导入类型: {import_type}")

    fields = IMPORT_FIELDS[import_type]
    wb = Workbook()
    ws = wb.active
    ws.title = f"{_type_name_cn(import_type)}导入模板"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    hint_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Row 1: Chinese headers
    headers = [hdr for _, hdr, _ in fields]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Row 2: example / hint row
    hints = _get_hint_row(import_type)
    for ci, h in enumerate(hints, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = hint_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + ci) if ci <= 26 else 'A'].width = max(14, min(len(h) * 2 + 4, 36))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ascii_name = f"{import_type}_template.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{ascii_name}"',
        },
    )


def _get_hint_row(import_type: str) -> list:
    """获取模板的示例行"""
    hints = {
        "member": [
            "张三", "13800138000", "男", "1990-01-01",
            "170", "65", "15", "普通", "正常",
            "门店推广", "S20250101001", "无", "",
        ],
        "staff": [
            "李四", "13900139000", "男", "1995-06-15",
            "教练", "2025-03-01", "5000", "在职", "", "", "30", "50",
        ],
        "membership_card": [
            "13800138000", "张三", "年卡", "黄金年卡", "365", "3000", "3000",
            "2025-01-01", "2025-12-31", "", "", "", "", "正常", "", "",
        ],
        "sold_lesson": [
            "13800138000", "张三", "减脂塑形课", "私教课", "2025-01-15", "24", "0", "24",
            "125", "3000", "3000", "微信", "13800138001", "",
        ],
        "course": [
            "减脂塑形课", "健身", "私教课", "24", "3000", "2500", "90", "上架", "10", "王教练",
        ],
        "checkin": [
            "13800138000", "张三", "2025-01-15", "08:30", "普通进场", "年卡", "", "刷卡", "", "前台小王",
        ],
        "body_measurement": [
            "13800138000", "张三", "2025-01-15", "170", "65", "15", "22.5", "55", "1500", "25",
        ],
    }
    return hints.get(import_type, [])


# ═══════════════════════════════════════════
# 上传解析 — POST /api/import/upload
# ═══════════════════════════════════════════

@router.post("/upload")
def upload_import_file(
    file: UploadFile = File(...),
    import_type: str = Form(...),
    request: Request = None,
    db: Session = Depends(get_db),
):
    """上传并解析导入文件，返回预览数据"""
    if import_type not in IMPORT_FIELDS:
        raise HTTPException(400, f"不支持的导入类型: {import_type}")

    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "请上传 .xlsx 格式的文件")

    # Read file
    content = file.file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"文件解析失败: {str(e)}")

    if ws.max_row < 2:
        raise HTTPException(400, "文件中没有数据（至少需要表头行 + 1行数据）")

    # Parse headers
    header_map = _build_header_map(import_type)
    raw_headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]

    # Map Chinese headers to field names
    col_mapping = {}  # column index (0-based) → field name
    for ci, h in enumerate(raw_headers):
        if h in header_map:
            col_mapping[ci] = header_map[h]

    if not col_mapping:
        raise HTTPException(400, "未能识别表头，请使用下载的模板填写")

    required_fields = {fld for fld, _, req in IMPORT_FIELDS[import_type] if req}
    mapped_fields = set(col_mapping.values())
    missing = required_fields - mapped_fields
    if missing:
        missing_names = [hdr for fld, hdr, _ in IMPORT_FIELDS[import_type] if fld in missing]
        raise HTTPException(400, f"缺少必填列: {', '.join(missing_names)}")

    # Parse data rows（先解析，暂不校验会员存在性）
    all_rows = []
    errors = []
    for ri in range(2, ws.max_row + 1):
        row_data = {}
        has_data = False
        for ci, field in col_mapping.items():
            raw_val = ws.cell(row=ri, column=ci + 1).value
            val = _parse_cell_value(raw_val, field, import_type)
            if val is not None:
                has_data = True
            row_data[field] = val
        if not has_data:
            continue  # skip empty rows

        # Validate required
        row_valid = True
        for fld in required_fields:
            val = row_data.get(fld)
            if val is None or (isinstance(val, str) and val.strip() == ""):
                errors.append({"row": ri, "field": fld, "message": f"「{dict(IMPORT_FIELDS[import_type])[fld]}」不能为空"})
                row_valid = False
        if not row_valid:
            continue

        all_rows.append(row_data)

    # 会员类导入：批量校验会员/员工存在性（消除 N+1 查询）
    if import_type in ("membership_card", "sold_lesson", "checkin", "body_measurement"):
        # 收集所有需要查询的手机号
        member_phones = set()
        staff_phones = set()
        for row in all_rows:
            p = str(row.get("member_phone") or "").strip()
            if p:
                member_phones.add(p)
            sp = str(row.get("staff_phone") or "").strip()
            if sp:
                staff_phones.add(sp)

        # 批量查询 Member
        member_map = {}
        if member_phones:
            for m in db.query(Member).filter(Member.phone.in_(member_phones)).all():
                member_map[m.phone] = m

        # 批量查询 Staff
        staff_map = {}
        if staff_phones:
            for s in db.query(Staff).filter(Staff.phone.in_(staff_phones)).all():
                staff_map[s.phone] = s

        # 逐行校验（内存操作，无 SQL 查询）
        valid_rows = []
        for i, row in enumerate(all_rows):
            ri = i + 2  # Excel 行号
            phone = str(row.get("member_phone") or "").strip()
            name = str(row.get("member_name") or "").strip()
            if phone:
                member = member_map.get(phone)
                if not member:
                    errors.append({"row": ri, "field": "member_phone", "message": f"手机号「{phone}」对应的会员不存在"})
                    continue
                if name and member.name != name:
                    errors.append({"row": ri, "field": "member_name", "message": f"会员姓名不匹配，系统记录为「{member.name}」"})
                    continue
                row["member_id"] = member.member_id
                if import_type == "membership_card":
                    row["member_name"] = member.name

            staff_phone = str(row.get("staff_phone") or "").strip()
            if staff_phone:
                staff = staff_map.get(staff_phone)
                if not staff:
                    errors.append({"row": ri, "field": "staff_phone", "message": f"销售员手机号「{staff_phone}」对应的员工不存在"})
                    continue
                row["staff_id"] = staff.staff_id
                row["staff_name"] = staff.name

            valid_rows.append(row)

        all_rows = valid_rows

    if not all_rows:
        raise HTTPException(400, "没有有效的数据行")

    # Create ImportTask
    operator = "系统"
    if request:
        token = request.cookies.get("access_token", "")
        if token:
            from jose import jwt
            from backend.routers.auth import SECRET_KEY, ALGORITHM
            try:
                payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
                operator = payload.get("sub", "系统")
            except Exception:
                pass

    task_id = generate_id("IT", db, ImportTask.task_id)
    preview = all_rows[:20]

    # 存储全部数据用于后台执行（JSON 序列化）
    task = ImportTask(
        task_id=task_id,
        import_type=import_type,
        status="confirming",
        total_rows=len(all_rows),
        preview_data=json.dumps(all_rows, ensure_ascii=False, default=str),
        filename=file.filename or "",
        operator=operator,
    )
    db.add(task)
    db.commit()

    # 构建字段名→中文名映射
    field_to_cn = {fld: hdr for fld, hdr, _ in IMPORT_FIELDS[import_type]}

    return {
        "task_id": task_id,
        "import_type": import_type,
        "total_rows": len(all_rows),
        "error_count": len(errors),
        "errors": errors,
        "preview": preview,
        "header_fields": list(col_mapping.values()),
        "header_names": [field_to_cn.get(f, f) for f in col_mapping.values()],
    }


# ═══════════════════════════════════════════
# 确认执行 — POST /api/import/{task_id}/execute
# ═══════════════════════════════════════════

@router.post("/{task_id}/execute")
def execute_import(task_id: str, db: Session = Depends(get_db)):
    """确认并开始执行导入（后台线程执行）"""
    task = db.query(ImportTask).filter(ImportTask.task_id == task_id).first()
    if not task:
        raise HTTPException(404, "导入任务不存在")
    if task.status not in ("confirming", "pending"):
        raise HTTPException(400, f"任务状态不正确（当前: {task.status}），无法执行")

    task.status = "pending"
    db.commit()

    # 使用 threading.Thread 确保后台可靠执行
    thread = threading.Thread(target=_run_import, args=(task_id, task.import_type), daemon=True)
    thread.start()

    return {"success": True, "message": "导入任务已启动", "task_id": task_id}


def _run_import(task_id: str, import_type: str):
    """后台线程执行导入（逐行 upsert）"""
    db = SessionLocal()
    try:
        task = db.query(ImportTask).filter(ImportTask.task_id == task_id).first()
        if not task:
            return

        task.status = "processing"
        db.commit()

        # 读取全部行数据
        rows_data = json.loads(task.preview_data or "[]")

        # 修复类型：JSON 反序列化后 date/Decimal 变成字符串
        date_fields = DATE_FIELDS.get(import_type, set())
        decimal_fields = DECIMAL_FIELDS.get(import_type, set())
        for row in rows_data:
            for fld in date_fields:
                v = row.get(fld)
                if isinstance(v, str) and v.strip():
                    try:
                        row[fld] = date.fromisoformat(v.strip())
                    except ValueError:
                        row[fld] = None
            for fld in decimal_fields:
                v = row.get(fld)
                if v is not None and v != '':
                    try:
                        row[fld] = Decimal(str(v))
                    except (ValueError, TypeError):
                        row[fld] = Decimal(0)

        config = TYPE_CONFIG.get(import_type)
        if not config:
            task.status = "failed"
            task.errors = json.dumps([{"row": 0, "message": f"未知导入类型: {import_type}"}])
            db.commit()
            return

        key_field = config["key_field"]
        protected = PROTECTED_FIELDS.get(import_type, set())

        total = len(rows_data)
        success = 0
        error_list = []
        created = 0
        updated = 0
        skipped = 0

        # 会员导入：严格防重复 — 批量校验 phone / name（消除 N+1 查询）
        if import_type == "member":
            phones_set = set()
            names_set = set()
            for row in rows_data:
                p = str(row.get("phone") or "").strip()
                if p:
                    phones_set.add(p)
                n = str(row.get("name") or "").strip()
                if n:
                    names_set.add(n)

            existing_phones = set()
            if phones_set:
                for m in db.query(Member.phone).filter(Member.phone.in_(phones_set)).all():
                    existing_phones.add(m[0])
            existing_names = set()
            if names_set:
                for m in db.query(Member.name).filter(Member.name.in_(names_set)).all():
                    existing_names.add(m[0])

            duplicate_found = False
            for row in rows_data:
                phone = str(row.get("phone") or "").strip()
                name = str(row.get("name") or "").strip()
                if phone and phone in existing_phones:
                    error_list.append({"row": 1, "message": f"手机号「{phone}」已存在，会员导入不允许重复"})
                    duplicate_found = True
                    break
                if name and name in existing_names:
                    error_list.append({"row": 1, "message": f"姓名「{name}」已存在，会员导入不允许重复"})
                    duplicate_found = True
                    break
            if duplicate_found:
                task.status = "failed"
                task.error_count = len(error_list)
                task.errors = json.dumps(error_list, ensure_ascii=False)
                db.commit()
                return

        # 预加载 staff / course / member 映射（消除 N+1 查询）
        existing_lookup = {}
        if import_type == "staff":
            for s in db.query(Staff).all():
                if s.phone:
                    existing_lookup[s.phone] = s
        elif import_type == "course":
            for c in db.query(Course).all():
                if c.name:
                    existing_lookup[c.name] = c

        member_cache = {}
        if import_type == "membership_card":
            member_ids = {row.get("member_id") for row in rows_data if row.get("member_id")}
            if member_ids:
                for m in db.query(Member).filter(Member.member_id.in_(member_ids)).all():
                    member_cache[m.member_id] = m

        for i, row in enumerate(rows_data):
            try:
                if key_field:
                    # UPSERT 模式（仅 staff 使用）
                    key_val = row.get(key_field)
                    if not key_val or (isinstance(key_val, str) and key_val.strip() == ""):
                        error_list.append({"row": i + 2, "field": key_field, "message": "匹配字段为空"})
                        skipped += 1
                        continue

                    key_val = str(key_val).strip()
                    existing = existing_lookup.get(key_val)

                    if existing:
                        _update_record(existing, row, key_field, protected)
                        updated += 1
                    else:
                        _create_record(db, import_type, row)
                        created += 1
                else:
                    # INSERT 模式（member / membership_card / sold_lesson）
                    _create_record(db, import_type, row)
                    created += 1

                success += 1

            except Exception as e:
                error_list.append({"row": i + 2, "message": str(e)})

            # 每 10 行更新进度
            if (i + 1) % 10 == 0 or i == total - 1:
                task.success_count = success
                task.error_count = len(error_list)
                task.created_count = created
                task.updated_count = updated
                task.skipped_count = skipped
                task.errors = json.dumps(error_list, ensure_ascii=False)
                db.commit()

        task.status = "completed"
        task.completed_at = datetime.now()
        task.success_count = success
        task.error_count = len(error_list)
        task.created_count = created
        task.updated_count = updated
        task.skipped_count = skipped
        task.errors = json.dumps(error_list, ensure_ascii=False)
        db.commit()

    except Exception as e:
        try:
            task = db.query(ImportTask).filter(ImportTask.task_id == task_id).first()
            if task:
                task.status = "failed"
                task.errors = json.dumps([{"row": 0, "message": str(e)}], ensure_ascii=False)
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


def _find_existing(db: Session, import_type: str, key_field: str, key_val: str):
    """查找已有记录（用于 upsert）"""
    if import_type == "staff":
        return db.query(Staff).filter(Staff.phone == key_val).first()
    return None


def _update_record(record, row: dict, key_field: str, protected: set):
    """更新已有记录的字段"""
    for field, val in row.items():
        if field == key_field or field in protected:
            continue
        if val is not None:
            setattr(record, field, val)


def _create_record(db: Session, import_type: str, row: dict, member_cache: dict = None):
    """创建新记录"""
    if import_type == "member":
        new_id = generate_id("M", db, Member.member_id)
        obj = Member(member_id=new_id)
    elif import_type == "staff":
        new_id = generate_id("S", db, Staff.staff_id)
        obj = Staff(staff_id=new_id)
    elif import_type == "membership_card":
        new_id = generate_id("MC", db, MembershipCard.card_id)
        obj = MembershipCard(card_id=new_id, is_product=0)
    elif import_type == "sold_lesson":
        new_id = generate_id("SL", db, Sale.sale_id)
        obj = Sale(sale_id=new_id)
    elif import_type == "course":
        new_id = generate_id("CO", db, Course.course_id)
        obj = Course(course_id=new_id)
    elif import_type == "checkin":
        new_id = generate_id("CI", db, Checkin.checkin_id)
        obj = Checkin(checkin_id=new_id)
    elif import_type == "body_measurement":
        new_id = generate_id("BM", db, BodyMeasurement.measure_id)
        obj = BodyMeasurement(measure_id=new_id)
    else:
        return

    # 仅跳过目标表中不存在的字段
    skip_fields = set()
    if import_type == "membership_card":
        skip_fields.update({"member_phone", "staff_phone"})
    if import_type == "sold_lesson":
        skip_fields.add("staff_phone")
    if import_type in ("checkin", "body_measurement"):
        skip_fields.add("member_phone")

    for field, val in row.items():
        if field in skip_fields:
            continue
        if val is not None:
            setattr(obj, field, val)
    db.add(obj)

    # 卡种联动：更新会员资产（使用缓存消除 N+1 查询）
    if import_type == "membership_card":
        card_type = str(row.get("card_type") or "").strip()
        member_id = row.get("member_id")
        if member_id:
            member = member_cache.get(member_id) if member_cache else None
            if member:
                if card_type in ("次卡",):
                    add_classes = (row.get("total_classes") or 0) + (row.get("bonus_classes") or 0)
                    if add_classes:
                        member.remaining_lessons = (member.remaining_lessons or 0) + add_classes
                elif card_type in ("现金卡", "现金"):
                    price_val = row.get("price") or 0
                    if price_val > 0:
                        member.balance = (member.balance or 0) + Decimal(str(price_val))


# ═══════════════════════════════════════════
# 进度查询 — GET /api/import/{task_id}/progress
# ═══════════════════════════════════════════

@router.get("/{task_id}/progress")
def get_import_progress(task_id: str, db: Session = Depends(get_db)):
    """查询导入进度"""
    task = db.query(ImportTask).filter(ImportTask.task_id == task_id).first()
    if not task:
        raise HTTPException(404, "导入任务不存在")

    return {
        "task_id": task.task_id,
        "import_type": task.import_type,
        "status": task.status,
        "total_rows": task.total_rows,
        "success_count": task.success_count,
        "error_count": task.error_count,
        "created_count": task.created_count,
        "updated_count": task.updated_count,
        "skipped_count": task.skipped_count,
        "errors": json.loads(task.errors or "[]"),
        "completed_at": task.completed_at.isoformat() if task.completed_at else None,
    }


# ═══════════════════════════════════════════
# SSE 实时进度推送
# ═══════════════════════════════════════════

@router.get("/{task_id}/progress/stream")
async def stream_import_progress(task_id: str):
    """SSE 实时进度推送（替代前端轮询）"""
    async def event_generator():
        while True:
            db = SessionLocal()
            try:
                task = db.query(ImportTask).filter(ImportTask.task_id == task_id).first()
                if not task:
                    yield f"event: error\ndata: {json.dumps({'message': '任务不存在'})}\n\n"
                    break

                data = {
                    "task_id": task.task_id,
                    "status": task.status,
                    "total_rows": task.total_rows,
                    "success_count": task.success_count,
                    "error_count": task.error_count,
                    "created_count": task.created_count,
                    "updated_count": task.updated_count,
                    "skipped_count": task.skipped_count,
                }

                yield f"event: progress\ndata: {json.dumps(data, ensure_ascii=False, default=str)}\n\n"

                if task.status in ("completed", "failed"):
                    data["errors"] = json.loads(task.errors or "[]")
                    data["completed_at"] = task.completed_at.isoformat() if task.completed_at else None
                    event_type = "complete" if task.status == "completed" else "error"
                    yield f"event: {event_type}\ndata: {json.dumps(data, ensure_ascii=False, default=str)}\n\n"
                    break
            except Exception as e:
                yield f"event: error\ndata: {json.dumps({'message': str(e)})}\n\n"
                break
            finally:
                db.close()

            await asyncio.sleep(0.5)

    return StreamingResponse(event_generator(), media_type="text/event-stream")


# ═══════════════════════════════════════════
# 导出导入结果 — GET /api/import/{task_id}/export
# ═══════════════════════════════════════════

@router.get("/{task_id}/export")
def export_import_result(task_id: str, db: Session = Depends(get_db)):
    """下载导入结果 XLSX（成功数据 + 失败数据 + 错误信息）"""
    task = db.query(ImportTask).filter(ImportTask.task_id == task_id).first()
    if not task:
        raise HTTPException(404, "导入任务不存在")

    all_rows = json.loads(task.preview_data or "[]")
    error_list = json.loads(task.errors or "[]")
    if not all_rows:
        raise HTTPException(400, "没有可导出的数据")

    # 构建字段→中文名映射
    field_to_cn = {fld: hdr for fld, hdr, _ in IMPORT_FIELDS.get(task.import_type, [])}

    # 提取失败行索引（error.row 是 Excel 行号 = 数据索引 + 2）
    failed_indices = set()
    row_error_map = {}
    batch_errors = []
    for err in error_list:
        r = err.get("row", 0)
        if r >= 2:
            idx = r - 2
            failed_indices.add(idx)
            row_error_map.setdefault(idx, []).append(err.get("message", ""))
        else:
            batch_errors.append(err.get("message", ""))

    # 分离成功/失败行
    success_rows = []
    failed_rows = []
    if task.status == "failed" and batch_errors:
        # 批处理失败（如会员重复）：所有行归入失败 sheet
        batch_msg = "；".join(batch_errors)
        for i, row in enumerate(all_rows):
            if i in failed_indices:
                failed_rows.append((row, "；".join(row_error_map[i])))
            else:
                failed_rows.append((row, batch_msg))
    else:
        for i, row in enumerate(all_rows):
            if i in failed_indices:
                failed_rows.append((row, "；".join(row_error_map.get(i, []))))
            else:
                success_rows.append(row)

    # 生成 XLSX
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    fields = [fld for fld, _, _ in IMPORT_FIELDS.get(task.import_type, [])]
    headers = [field_to_cn.get(f, f) for f in fields]

    def _write_sheet(ws, title, data_rows, extra_col=None):
        ws.title = title
        cols = headers + ([extra_col] if extra_col else [])
        for ci, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for ri, row_data in enumerate(data_rows, 2):
            if isinstance(row_data, tuple):
                row_dict, extra_val = row_data
            else:
                row_dict, extra_val = row_data, None

            for ci, fld in enumerate(fields, 1):
                val = row_dict.get(fld)
                if val is None:
                    val = ""
                elif isinstance(val, (date, datetime)):
                    val = val.isoformat()
                ws.cell(row=ri, column=ci, value=val)

            if extra_col and extra_val is not None:
                ws.cell(row=ri, column=len(cols), value=extra_val)

        for ci, h in enumerate(cols, 1):
            ws.column_dimensions[chr(64 + ci) if ci <= 26 else 'A'].width = max(12, min(len(h) * 2 + 4, 40))

    # Sheet 1: 成功导入
    ws_success = wb.active
    _write_sheet(ws_success, "成功导入", success_rows)

    # Sheet 2: 导入失败
    if failed_rows:
        ws_fail = wb.create_sheet()
        _write_sheet(ws_fail, "导入失败", failed_rows, extra_col="错误信息")
    else:
        ws_fail = wb.create_sheet()
        ws_fail.title = "导入失败"
        for ci, h in enumerate(headers + ["错误信息"], 1):
            cell = ws_fail.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
        ws_fail.cell(row=2, column=1, value="无失败记录")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ascii_name = f"import_result_{task_id}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{ascii_name}"',
        },
    )


# ═══════════════════════════════════════════
# 导入历史 — HTMX 表格
# ═══════════════════════════════════════════

TYPE_CN = {"member": "会员", "staff": "员工", "membership_card": "会籍卡", "sold_lesson": "私教课",
           "course": "课程", "checkin": "进场记录", "body_measurement": "体测记录"}
STATUS_CN = {
    "pending": "等待中", "parsing": "解析中", "confirming": "待确认",
    "processing": "导入中", "completed": "已完成", "failed": "失败",
}
STATUS_CLS = {
    "completed": "text-green-600", "failed": "text-red-600",
    "processing": "text-blue-600", "pending": "text-gray-500",
    "confirming": "text-yellow-600",
}


@router.get("/history/table", response_class=HTMLResponse)
def import_history_table(db: Session = Depends(get_db)):
    """导入历史 HTML 表格"""
    tasks = db.query(ImportTask).order_by(ImportTask.created_at.desc()).limit(50).all()
    if not tasks:
        return '<div class="text-center py-8 text-gray-400">暂无导入记录</div>'

    rows = ""
    for t in tasks:
        status_cn = STATUS_CN.get(t.status, t.status)
        status_cls = STATUS_CLS.get(t.status, "text-gray-500")
        type_cn = TYPE_CN.get(t.import_type, t.import_type)
        completed = t.completed_at.strftime('%Y-%m-%d %H:%M') if t.completed_at else '-'

        # 操作按钮：待确认 → 可执行；已完成/失败 → 可下载
        if t.status == "confirming":
            action_btn = f"""<button class="text-blue-600 hover:text-blue-800 text-xs"
                hx-post="/api/import/{t.task_id}/execute"
                hx-swap="none"
                hx-on::after-request="if(event.detail.successful){{ htmx.ajax('GET','/api/import/history/table',{{target:'#historyTable',swap:'innerHTML'}}); alert('导入任务已启动'); }}">
                执行导入</button>"""
        elif t.status in ("completed", "failed") and t.preview_data:
            action_btn = f"""<a href="/api/import/{t.task_id}/export"
                class="text-green-600 hover:text-green-800 text-xs underline">下载结果</a>"""
        else:
            action_btn = '<span class="text-gray-400 text-xs">-</span>'

        rows += f"""<tr class="hover:bg-gray-50 border-b">
            <td class="px-4 py-3 text-sm">{t.task_id}</td>
            <td class="px-4 py-3 text-sm">{type_cn}</td>
            <td class="px-4 py-3 text-sm">{t.filename}</td>
            <td class="px-4 py-3 text-sm">{t.total_rows}</td>
            <td class="px-4 py-3 text-sm text-green-600">{t.success_count}</td>
            <td class="px-4 py-3 text-sm text-red-600">{t.error_count}</td>
            <td class="px-4 py-3 text-sm text-blue-600">{t.created_count}</td>
            <td class="px-4 py-3 text-sm text-orange-600">{t.updated_count}</td>
            <td class="px-4 py-3 text-sm {status_cls} font-medium">{status_cn}</td>
            <td class="px-4 py-3 text-sm">{completed}</td>
            <td class="px-4 py-3 text-sm">{t.operator}</td>
            <td class="px-4 py-3 text-sm">{action_btn}</td>
        </tr>"""

    return f"""<table class="w-full bg-white rounded-lg shadow-sm">
        <thead class="bg-gray-50 text-left text-xs text-gray-500 uppercase">
            <tr><th class="px-4 py-3">任务编号</th><th class="px-4 py-3">类型</th><th class="px-4 py-3">文件</th>
                <th class="px-4 py-3">总行数</th><th class="px-4 py-3">成功</th><th class="px-4 py-3">失败</th>
                <th class="px-4 py-3">新建</th><th class="px-4 py-3">更新</th><th class="px-4 py-3">状态</th>
                <th class="px-4 py-3">完成时间</th><th class="px-4 py-3">操作人</th><th class="px-4 py-3">操作</th></tr>
        </thead>
        <tbody>{rows}</tbody>
    </table>"""


@router.get("/history")
def import_history_list(skip: int = 0, limit: int = 50, db: Session = Depends(get_db)):
    """导入历史 JSON 列表"""
    tasks = db.query(ImportTask).order_by(ImportTask.created_at.desc()).offset(skip).limit(limit).all()
    return [
        {
            "task_id": t.task_id,
            "import_type": t.import_type,
            "status": t.status,
            "total_rows": t.total_rows,
            "success_count": t.success_count,
            "error_count": t.error_count,
            "created_count": t.created_count,
            "updated_count": t.updated_count,
            "filename": t.filename,
            "operator": t.operator,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "completed_at": t.completed_at.isoformat() if t.completed_at else None,
        }
        for t in tasks
    ]
