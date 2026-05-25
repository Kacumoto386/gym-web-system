# -*- coding: utf-8 -*-
"""
鹤壁飞鸟 Excel 数据导入工具
==========================
从 data/鹤壁飞鸟1.0.xlsx 导入可用数据到系统数据库。

用法:
    python scripts/import_excel_data.py

设计:
    - 只导入有效/可用的数据（跳过空行、无效关联）
    - 幂等：可重复运行，已存在的记录自动跳过
    - 通过手机号关联会员（因原系统会员卡号多为空）
    - 大表批量提交（每 500 行 flush）
"""
import sys
import os
from datetime import datetime, date, timedelta

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from backend.database import SessionLocal, init_db
from backend.models.models import (
    Staff, Member, MembershipCard, Sale, LessonPackage,
    ClassRecord, Checkin, Product, ProductSale, Course, Store,
)
from sqlalchemy import func


# ═══════════════════════════════════════════
# 安全类型转换
# ═══════════════════════════════════════════

def safe_int(v, default=0):
    if v is None or str(v).strip() in ("", "-"):
        return default
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return default


def safe_float(v, default=0):
    if v is None or str(v).strip() in ("", "-"):
        return default
    try:
        return float(str(v))
    except (ValueError, TypeError):
        return default


def safe_str(v, default=""):
    if v is None:
        return default
    return str(v).strip()


def safe_date(v):
    """支持多种日期格式的解析"""
    if v is None or str(v).strip() in ("", "-"):
        return None
    v = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(v, fmt).date()
        except ValueError:
            continue
    return None


def safe_datetime(v):
    """解析日期时间字符串"""
    if v is None or str(v).strip() in ("", "-"):
        return None
    v = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(v, fmt)
        except ValueError:
            continue
    return None


def extract_time(v):
    """从日期时间字符串中提取时间部分 (HH:MM)"""
    if v is None or str(v).strip() in ("", "-"):
        return ""
    v = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            dt = datetime.strptime(v, fmt)
            return dt.strftime("%H:%M")
        except ValueError:
            continue
    return ""


# ═══════════════════════════════════════════
# ID 生成器（内存计数器，避免每次查库）
# ═══════════════════════════════════════════

class IDGenerator:
    """内存 ID 生成器，格式: {PREFIX}{YYYYMMDD}{NNNN}"""
    def __init__(self, db, prefix, column):
        self.prefix = prefix
        self.date_str = date.today().strftime("%Y%m%d")
        # 查库获取当前最大序号
        max_id = db.query(func.max(column)).filter(
            column.like(f"{prefix}{self.date_str}%")
        ).scalar()
        if max_id:
            self.seq = int(str(max_id)[-4:]) + 1
        else:
            self.seq = 1

    def next(self):
        n = self.seq
        self.seq += 1
        return f"{self.prefix}{self.date_str}{n:04d}"


# ═══════════════════════════════════════════
# Excel 数据加载
# ═══════════════════════════════════════════

def load_sheet_by_index(wb, index):
    """按索引加载 Sheet 数据，返回 header→value 的字典列表"""
    ws = wb[wb.sheetnames[index]]
    headers = []
    for row in ws.iter_rows(max_row=1, values_only=True):
        for h in row:
            headers.append(safe_str(h))
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row_data = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                row_data[headers[j]] = val
        # 检查是否为空行（所有字段为空）
        if any(v is not None and str(v).strip() for v in row if v is not None):
            rows.append(row_data)
    return rows


# ═══════════════════════════════════════════
# 幂等性检查
# ═══════════════════════════════════════════

def exists_staff(db, phone):
    return db.query(Staff).filter(Staff.phone == phone).first() is not None


def exists_member(db, phone):
    return db.query(Member).filter(Member.phone == phone).first() is not None


def exists_card(db, member_id, card_type, start_date, price):
    q = db.query(MembershipCard).filter(
        MembershipCard.member_id == member_id,
        MembershipCard.card_type == card_type,
    )
    if start_date:
        q = q.filter(MembershipCard.start_date == start_date)
    if price:
        q = q.filter(MembershipCard.price == price)
    return q.first() is not None


def exists_sale(db, member_id, course_name, sale_date):
    q = db.query(Sale).filter(
        Sale.member_id == member_id,
        Sale.course_name == course_name,
    )
    if sale_date:
        q = q.filter(Sale.sale_date == sale_date)
    return q.first() is not None


def exists_class_record(db, member_id, class_date, course_name):
    q = db.query(ClassRecord).filter(
        ClassRecord.member_id == member_id,
        ClassRecord.class_date == class_date,
        ClassRecord.course_name == course_name,
    )
    return q.first() is not None


def exists_checkin(db, member_id, checkin_date, checkin_time):
    q = db.query(Checkin).filter(
        Checkin.member_id == member_id,
        Checkin.checkin_date == checkin_date,
    )
    if checkin_time:
        q = q.filter(Checkin.checkin_time == checkin_time)
    return q.first() is not None


def exists_product(db, name):
    return db.query(Product).filter(Product.name == name).first() is not None


def exists_product_sale(db, product_name, sale_date, quantity):
    q = db.query(ProductSale).filter(
        ProductSale.product_name == product_name,
        ProductSale.sale_date == sale_date,
        ProductSale.quantity == quantity,
    )
    return q.first() is not None


def exists_store(db, store_id):
    return db.query(Store).filter(Store.store_id == store_id).first() is not None


def exists_course(db, name):
    return db.query(Course).filter(Course.name == name).first() is not None


# ═══════════════════════════════════════════
# 导入函数
# ═══════════════════════════════════════════

def import_store(db):
    """创建默认门店"""
    if exists_store(db, "HB001"):
        return 0
    store = Store(
        store_id="HB001",
        name="鹤壁店",
        address="鹤壁",
        status="正常",
    )
    db.add(store)
    db.commit()
    return 1


def import_staff(db, rows):
    """导入员工表 → Staff"""
    id_gen = IDGenerator(db, "S", Staff.staff_id)
    imported = 0
    skipped = 0
    for row in rows:
        name = safe_str(row.get("中文名"))
        if not name:
            skipped += 1
            continue
        phone = safe_str(row.get("手机号"))
        if exists_staff(db, phone):
            skipped += 1
            continue
        staff = Staff(
            staff_id=id_gen.next(),
            name=name,
            gender=safe_str(row.get("性别")),
            phone=phone,
            id_card=safe_str(row.get("证件号")),
            birth_date=safe_date(row.get("生日")),
            hire_date=safe_date(row.get("入职时间")),
            position=safe_str(row.get("职位")),
            status="在职" if safe_str(row.get("状态")) == "在职" else "离职",
            store_id="HB001",
        )
        db.add(staff)
        imported += 1
        if imported % 100 == 0:
            db.flush()
    db.commit()
    return imported, skipped


def import_courses(db, private_lesson_rows, class_record_rows):
    """从课程数据中提取课程列表 → Course"""
    names = set()
    # 从私教课程表提取
    for row in private_lesson_rows:
        t = safe_str(row.get("课程类型"))
        n = safe_str(row.get("课程名称"))
        if n:
            names.add(n if not t else f"{t}-{n}")
    # 从消费记录提取
    for row in class_record_rows:
        n = safe_str(row.get("课程名称"))
        if n:
            names.add(n)
    id_gen = IDGenerator(db, "C", Course.course_id)
    imported = 0
    for name in sorted(names):
        if not exists_course(db, name):
            course = Course(
                course_id=id_gen.next(),
                name=name,
                status="上架",
            )
            db.add(course)
            imported += 1
    db.commit()
    return imported, len(names) - imported


def import_members(db, rows):
    """导入会员信息 → Member，返回 phone_map 和 name_map"""
    id_gen = IDGenerator(db, "M", Member.member_id)
    phone_map = {}  # phone → member_id
    name_map = {}   # name → [member_id, ...]
    imported = 0
    skipped = 0

    # 状态映射
    status_map = {
        "历史会员": "已过期",
        "新客户": "正常",
        "正式会员": "正常",
    }

    for row in rows:
        name = safe_str(row.get("会员姓名"))
        phone = safe_str(row.get("手机号"))
        if not name or not phone:
            skipped += 1
            continue
        if exists_member(db, phone):
            skipped += 1
            continue

        raw_status = safe_str(row.get("会员状态"))
        status = status_map.get(raw_status, "正常")

        member = Member(
            member_id=id_gen.next(),
            name=name,
            gender=safe_str(row.get("性别")),
            phone=phone,
            id_card=safe_str(row.get("证件号")),
            birth_date=safe_date(row.get("生日")),
            source=safe_str(row.get("信息来源")),
            status=status,
            level="普通",
            staff_name=safe_str(row.get("会籍顾问")),
            start_date=safe_date(row.get("入会时间")),
            remark=safe_str(row.get("客户分类")),
            store_id="HB001",
        )
        db.add(member)
        db.flush()
        imported += 1

        # 构建映射缓存
        phone_map[phone] = member.member_id
        if name not in name_map:
            name_map[name] = []
        name_map[name].append(member.member_id)

        if imported % 100 == 0:
            db.flush()

    db.commit()
    return imported, skipped, phone_map, name_map


def find_member_id(name, phone, phone_map, name_map):
    """通过手机号/姓名查找会员ID"""
    if phone and phone in phone_map:
        return phone_map[phone]
    if name and name in name_map:
        ids = name_map[name]
        if len(ids) == 1:
            return ids[0]
    return None


def import_membership_cards(db, rows, phone_map, name_map):
    """导入会籍购买表 → MembershipCard"""
    id_gen = IDGenerator(db, "MC", MembershipCard.card_id)
    imported = 0
    skipped = 0
    for row in rows:
        name = safe_str(row.get("会员名称"))
        phone = safe_str(row.get("会员手机号"))
        if not name:
            skipped += 1
            continue
        member_id = find_member_id(name, phone, phone_map, name_map)
        if not member_id:
            skipped += 1
            continue
        card_type = safe_str(row.get("会籍类型"))
        start_date = safe_date(row.get("开始时间"))
        end_date = safe_date(row.get("结束时间"))
        price = safe_float(row.get("应收"))
        if exists_card(db, member_id, card_type, start_date, price):
            skipped += 1
            continue
        card = MembershipCard(
            card_id=id_gen.next(),
            member_id=member_id,
            member_name=name,
            card_type=card_type,
            duration_days=safe_int(row.get("有效期")),
            start_date=start_date,
            end_date=end_date,
            price=price,
            status="已过期" if safe_str(row.get("卡状态")) == "已过期" else safe_str(row.get("卡状态"), "正常"),
            remark=safe_str(row.get("备注")),
            store_id="HB001",
        )
        db.add(card)
        imported += 1
        if imported % 100 == 0:
            db.flush()
    db.commit()
    return imported, skipped


def import_sales_and_packages(db, rows, phone_map, name_map):
    """导入私教课程表 → Sale + LessonPackage"""
    sale_id_gen = IDGenerator(db, "SL", Sale.sale_id)
    pkg_id_gen = IDGenerator(db, "LP", LessonPackage.package_id)
    sale_imported = 0
    pkg_imported = 0
    skipped = 0
    for row in rows:
        name = safe_str(row.get("会员名称"))
        phone = safe_str(row.get("手机号"))
        if not name:
            skipped += 1
            continue
        member_id = find_member_id(name, phone, phone_map, name_map)
        if not member_id:
            skipped += 1
            continue
        course_type = safe_str(row.get("课程类型"))
        course_name = safe_str(row.get("课程名称"))
        full_name = f"{course_type}-{course_name}" if course_type else course_name
        sale_date = safe_date(row.get("购买时间"))
        total_hours = safe_int(row.get("购买课数量"))

        if not full_name or not sale_date:
            skipped += 1
            continue

        # Sale
        sale = None
        if not exists_sale(db, member_id, full_name, sale_date):
            sale_id = sale_id_gen.next()
            sale = Sale(
                sale_id=sale_id,
                member_id=member_id,
                member_name=name,
                course_name=full_name,
                sale_date=sale_date,
                total_hours=total_hours,
                total_price=safe_float(row.get("应收总价")),
                actual_amount=safe_float(row.get("实收总价")),
                staff_name=safe_str(row.get("销售员")),
                start_date=safe_date(row.get("开始日期")),
                end_date=safe_date(row.get("结束日期")),
                payment_method=safe_str(row.get("销售方式")),
                payment_status="已付清",
                store_id="HB001",
            )
            db.add(sale)
            sale_imported += 1
        else:
            # 已存在 sale，查找它用于关联 package
            existing = db.query(Sale).filter(
                Sale.member_id == member_id,
                Sale.course_name == full_name,
                Sale.sale_date == sale_date,
            ).first()
            if existing:
                sale = existing

        # LessonPackage
        pkg_key = (member_id, full_name, total_hours)
        pkg_exists = db.query(LessonPackage).filter(
            LessonPackage.member_id == member_id,
            LessonPackage.course_name == full_name,
            LessonPackage.total_hours == total_hours,
        ).first() is not None

        if not pkg_exists:
            remaining = safe_int(row.get("剩余购买课"))
            used = safe_int(row.get("总上课数量"))
            pkg = LessonPackage(
                package_id=pkg_id_gen.next(),
                sale_id=sale.sale_id if sale else "",
                member_id=member_id,
                member_name=name,
                course_name=full_name,
                total_hours=total_hours,
                remaining_hours=remaining,
                used_hours=used,
                valid_from=safe_date(row.get("开始日期")),
                valid_until=safe_date(row.get("结束日期")),
                status="正常" if remaining > 0 else "已完成",
                store_id="HB001",
            )
            db.add(pkg)
            pkg_imported += 1

        if (sale_imported + pkg_imported) % 100 == 0:
            db.flush()
    db.commit()
    return sale_imported, pkg_imported, skipped


def import_class_records(db, rows, phone_map, name_map):
    """导入私教课消费记录 → ClassRecord"""
    id_gen = IDGenerator(db, "CL", ClassRecord.record_id)
    imported = 0
    skipped = 0
    for row in rows:
        name = safe_str(row.get("会员姓名"))
        phone = safe_str(row.get("会员手机号"))
        if not name:
            skipped += 1
            continue
        member_id = find_member_id(name, phone, phone_map, name_map)
        if not member_id:
            skipped += 1
            continue
        class_date = safe_date(row.get("预约上课时间"))
        course_name = safe_str(row.get("课程名称"))
        if not class_date or not course_name:
            skipped += 1
            continue
        if exists_class_record(db, member_id, class_date, course_name):
            skipped += 1
            continue
        record = ClassRecord(
            record_id=id_gen.next(),
            member_id=member_id,
            member_name=name,
            course_name=course_name,
            class_date=class_date,
            start_time=extract_time(row.get("预约上课时间")),
            consumed_hours=safe_int(row.get("课时"), 1),
            coach_name=safe_str(row.get("上课教练")),
            sign_in_time=safe_datetime(row.get("确认消费时间")),
            status="已完成",
            sign_in_count=1,
            store_id="HB001",
        )
        db.add(record)
        imported += 1
        if imported % 100 == 0:
            db.flush()
    db.commit()
    return imported, skipped


def import_checkins(db, rows, phone_map, name_map):
    """导入进店记录 → Checkin（批量提交）"""
    id_gen = IDGenerator(db, "CI", Checkin.checkin_id)
    imported = 0
    skipped = 0
    batch = []
    for i, row in enumerate(rows):
        name = safe_str(row.get("会员姓名"))
        phone = safe_str(row.get("会员手机号"))
        if not name:
            skipped += 1
            continue
        member_id = find_member_id(name, phone, phone_map, name_map)
        if not member_id:
            skipped += 1
            continue
        checkin_date = safe_date(row.get("进店时间"))
        checkin_time = extract_time(row.get("进店时间"))
        if not checkin_date:
            skipped += 1
            continue
        if exists_checkin(db, member_id, checkin_date, checkin_time):
            skipped += 1
            continue
        ci = Checkin(
            checkin_id=id_gen.next(),
            member_id=member_id,
            member_name=name,
            checkin_date=checkin_date,
            checkin_time=checkin_time,
            card_type=safe_str(row.get("会籍卡")),
            operator=safe_str(row.get("进店操作人")),
            store_id="HB001",
        )
        db.add(ci)
        imported += 1
        if imported % 500 == 0:
            db.flush()
    db.commit()
    return imported, skipped


def import_products(db, rows):
    """导入商品库存 → Product"""
    id_gen = IDGenerator(db, "P", Product.product_id)
    imported = 0
    skipped = 0
    for row in rows:
        name = safe_str(row.get("商品名称"))
        if not name:
            skipped += 1
            continue
        if exists_product(db, name):
            skipped += 1
            continue
        brand = safe_str(row.get("品牌"))
        remark = f"品牌:{brand}" if brand else ""
        product = Product(
            product_id=id_gen.next(),
            name=name,
            category=safe_str(row.get("商品分类")),
            cost_price=safe_float(row.get("成本价")),
            selling_price=safe_float(row.get("价格")),
            stock=safe_int(row.get("库存数量")),
            unit="个",
            remark=remark,
        )
        db.add(product)
        imported += 1
    db.commit()
    return imported, skipped


def import_product_sales(db, rows, phone_map, name_map):
    """导入商品购买记录 → ProductSale"""
    id_gen = IDGenerator(db, "PS", ProductSale.sale_id)
    imported = 0
    skipped = 0
    for row in rows:
        product_name = safe_str(row.get("商品名称"))
        sale_date = safe_date(row.get("购买时间"))
        quantity = safe_int(row.get("购买数量"), 1)
        if not product_name or not sale_date:
            skipped += 1
            continue
        if exists_product_sale(db, product_name, sale_date, quantity):
            skipped += 1
            continue

        # 尝试关联会员（可选）
        name = safe_str(row.get("会员姓名"))
        phone = safe_str(row.get("会员卡号"), "")
        member_id = find_member_id(name, phone, phone_map, name_map) if name else None

        ps = ProductSale(
            sale_id=id_gen.next(),
            sale_date=sale_date,
            product_name=product_name,
            quantity=quantity,
            unit_price=safe_float(row.get("应收额")) / max(quantity, 1),
            total_price=safe_float(row.get("应收额")),
            payment_method=safe_str(row.get("支付方式")),
            member_id=member_id or "",
            member_name=name,
            operator=safe_str(row.get("销售员")),
        )
        db.add(ps)
        imported += 1
        if imported % 100 == 0:
            db.flush()
    db.commit()
    return imported, skipped


# ═══════════════════════════════════════════
# 主流程
# ═══════════════════════════════════════════

def main():
    EXCEL_PATH = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "data", "鹤壁飞鸟1.0.xlsx"
    )
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel 文件不存在: {EXCEL_PATH}")
        return

    print(f"[Excel] 加载: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"  共 {len(wb.sheetnames)} 个 Sheet")

    # 加载各 Sheet 数据
    print("\n[Data] 读取数据...")
    sheet_names = wb.sheetnames
    for i, n in enumerate(sheet_names):
        print(f"  Sheet {i}: {n}")

    staff_rows = load_sheet_by_index(wb, 0)       # 员工表
    member_rows = load_sheet_by_index(wb, 2)       # 会员信息
    card_rows = load_sheet_by_index(wb, 3)         # 会籍购买表
    private_lesson_rows = load_sheet_by_index(wb, 4)  # 私教课程表
    class_record_rows = load_sheet_by_index(wb, 12)   # 私教课消费记录
    checkin_rows = load_sheet_by_index(wb, 14)        # 进店记录
    product_rows = load_sheet_by_index(wb, 16)        # 商品库存
    product_sale_rows = load_sheet_by_index(wb, 17)   # 商品购买记录
    wb.close()

    print(f"  员工表: {len(staff_rows)} 行")
    print(f"  会员信息: {len(member_rows)} 行")
    print(f"  会籍购买表: {len(card_rows)} 行")
    print(f"  私教课程表: {len(private_lesson_rows)} 行")
    print(f"  私教课消费记录: {len(class_record_rows)} 行")
    print(f"  进店记录: {len(checkin_rows)} 行")
    print(f"  商品库存: {len(product_rows)} 行")
    print(f"  商品购买记录: {len(product_sale_rows)} 行")

    # 数据库初始化
    init_db()
    db = SessionLocal()

    results = {}

    try:
        # ── Step 1: Store ──
        print("\n[Store] 门店...")
        results["store"] = import_store(db)
        print(f"  → 导入: {results['store']}")

        # ── Step 2: Staff ──
        print("\n[Staff] 员工...")
        imp, skp = import_staff(db, staff_rows)
        results["staff"] = (imp, skp)
        print(f"  → 导入: {imp}, 跳过: {skp}")

        # ── Step 3: Course ──
        print("\n[Course] 课程...")
        imp, skp = import_courses(db, private_lesson_rows, class_record_rows)
        results["course"] = (imp, skp)
        print(f"  → 导入: {imp}, 跳过（已存在）: {skp}")

        # ── Step 4: Member ──
        print("\n[Member] 会员...")
        imp, skp, phone_map, name_map = import_members(db, member_rows)
        results["member"] = (imp, skp)
        print(f"  → 导入: {imp}, 跳过: {skp}")
        print(f"  → 映射: {len(phone_map)} 手机号, {len(name_map)} 姓名")

        # ── Step 5: MembershipCard ──
        print("\n[Card] 会籍卡...")
        imp, skp = import_membership_cards(db, card_rows, phone_map, name_map)
        results["membership_card"] = (imp, skp)
        print(f"  → 导入: {imp}, 跳过: {skp}")

        # ── Step 6: Sale + LessonPackage ──
        print("\n[Sale] 售课 + 课程包...")
        s_imp, p_imp, skp = import_sales_and_packages(db, private_lesson_rows, phone_map, name_map)
        results["sale"] = (s_imp, skp)
        results["lesson_package"] = (p_imp, 0)
        print(f"  → Sale 导入: {s_imp}, Package 导入: {p_imp}, 跳过: {skp}")

        # ── Step 7: ClassRecord ──
        print("\n[ClassRecord] 上课记录...")
        imp, skp = import_class_records(db, class_record_rows, phone_map, name_map)
        results["class_record"] = (imp, skp)
        print(f"  → 导入: {imp}, 跳过: {skp}")

        # ── Step 8: Checkin ──
        print("\n[Checkin] 进场记录...")
        imp, skp = import_checkins(db, checkin_rows, phone_map, name_map)
        results["checkin"] = (imp, skp)
        print(f"  → 导入: {imp}, 跳过: {skp}")

        # ── Step 9: Product ──
        print("\n[Product] 商品...")
        imp, skp = import_products(db, product_rows)
        results["product"] = (imp, skp)
        print(f"  → 导入: {imp}, 跳过: {skp}")

        # ── Step 10: ProductSale ──
        print("\n[ProductSale] 商品零售...")
        imp, skp = import_product_sales(db, product_sale_rows, phone_map, name_map)
        results["product_sale"] = (imp, skp)
        print(f"  → 导入: {imp}, 跳过: {skp}")

    except Exception as e:
        db.rollback()
        print(f"\n[ERROR] 导入失败: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()

    # ── 汇总 ──
    print("\n" + "=" * 60)
    print("  导入报告")
    print("=" * 60)
    total_imp = 0
    total_skp = 0
    for key, val in results.items():
        if isinstance(val, tuple):
            imp, skp = val
            print(f"  {key:20s}: {imp:>5d} 导入, {skp:>5d} 跳过")
            total_imp += imp
            total_skp += skp
        else:
            print(f"  {key:20s}: {val:>5d}")
    print("-" * 60)
    print(f"  {'总计':20s}: {total_imp:>5d} 导入, {total_skp:>5d} 跳过")
    print("=" * 60)


if __name__ == "__main__":
    main()
